#IMPORTS PARA GENERAR EL EXCEL
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,Font,Border,Side
import datetime

def GenerarExcel_3(datos):
    # #GENERANDO UN EXCEL CON LA INFORMACION DE ODDO
    # Crea un libro de trabajo y una hoja de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

   

    # Crea un objeto Font y ajusta la propiedad bold a True para establecer el texto en negritas
    fuente = Font(bold=True)
    

    # Crea un objeto Alignment y ajusta la propiedad vertical a 'top'
    alineacion = Alignment(horizontal='center', vertical='center', wrap_text=True)
    

    # Asignamos bordes a la celda
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Une las tres celdas para crear una celda combinada
   


   



    # Añade los nombres y apellidos a la hoja de trabajo en dos columnas separadas
    ws['A1'] = 'id_producto'
    ws['B1'] = 'producto'
    ws['C1'] = 'precio'

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20

    ws.row_dimensions[2].height = 15
    ws.row_dimensions[1].height = 11

    columnas = ['A','B','C']
    #APLICAMOS EL FORMATO A LAS CELDAS DEL FOR ENCABEZADOS

    for col in columnas:
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[col+'1'].border = border
        ws[col+'1'].alignment = alineacion
        ws[col+'1'].font = fuente
        


    filaCont = 2
    
    print("DATOS RECIBIDOS",datos)
    for fila in datos[0]:
        ws.cell(row=filaCont, column=1, value=fila[0])
        ws.cell(row=filaCont, column=2, value=fila[1])
        ws.cell(row=filaCont, column=3, value=fila[2])
        filaCont += 1
        
    
    fecha_hora_actual = datetime.datetime.now()
    fecha_hora_actual_formateada = fecha_hora_actual.strftime("%d-%m-%y_%H%M%S")
    

    
    nombre = "static/sistema/reportes/plantilla/Reporte_"+str(fecha_hora_actual_formateada)+".xlsx"
    wb.save(nombre)
    return nombre


def GenerarExcel_1(datos):
    # #GENERANDO UN EXCEL CON LA INFORMACION DE ODDO
    # Crea un libro de trabajo y una hoja de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Añade el texto en la celda A1
    ws['A1'] = 'COMPAÑÍA RECICLADORA DE NICARAGUA - REPORTE DE MANTENIMIENTO'

    # Obtiene la celda A1
    celda = ws['A1']


    # Crea un objeto Font y ajusta la propiedad bold a True para establecer el texto en negritas
    fuente = Font(bold=True)
    celda.font = fuente

    # Crea un objeto Alignment y ajusta la propiedad vertical a 'top'
    alineacion = Alignment(horizontal='center', vertical='center', wrap_text=True)
    celda.alignment = alineacion

    # Asignamos bordes a la celda
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    celda.border = border

    # Une las tres celdas para crear una celda combinada
    ws.merge_cells('A1:H1')


    # Inserta la imagen en la celda combinada B2:D4
    img = Image('static/img/LOGO CRN.png')
    img.width = 50
    img.height = 50
    ws.add_image(img, 'A1')



    # Añade los nombres y apellidos a la hoja de trabajo en dos columnas separadas
    ws['A2'] = 'Num'
    ws['B2'] = 'Máquina'
    ws['C2'] = 'Fecha Inicio'
    ws['D2'] = 'Fecha Fin'
    ws['E2'] = 'Tipo'
    ws['F2'] = 'Asignado'
    ws['G2'] = 'Costo'
    ws['H2'] = 'Estado'

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 11

    ws.row_dimensions[2].height = 15
    ws.row_dimensions[1].height = 44

    columnas = ['A','B','C','D','E','F','G','H']
    #APLICAMOS EL FORMATO A LAS CELDAS DEL FOR ENCABEZADOS

    for col in columnas:
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[col+'2'].border = border
        ws[col+'2'].alignment = alineacion
        ws[col+'2'].font = fuente
        


    filaCont = 3
    contador = 0
    print("DATOS RECIBIDOS",datos[0])
    for fila in datos:
        ws.cell(row=filaCont, column=1, value=fila[0][0])
        ws.cell(row=filaCont, column=2, value=fila[0][1])
        ws.cell(row=filaCont, column=3, value=fila[0][2])
        ws.cell(row=filaCont, column=4, value=fila[0][3])
        ws.cell(row=filaCont, column=5, value=fila[0][4])
        ws.cell(row=filaCont, column=6, value=fila[0][5])
        ws.cell(row=filaCont, column=7, value='C$ '+ str(fila[0][6]))
        ws.cell(row=filaCont, column=8, value=fila[0][7])
        filaCont += 1
        contador += 1
        
    
    fecha_hora_actual = datetime.datetime.now()
    fecha_hora_actual_formateada = fecha_hora_actual.strftime("%d-%m-%y_%H%M%S")
    

    
    # output = BytesIO()
    # wb.save(output)
    # output.seek(0)
    nombre = "static/Reportes/Reporte_"+str(fecha_hora_actual_formateada)+".xlsx"
    wb.save(nombre)
    return nombre
# controllers/web.py
import os
from flask import Blueprint
import random
import string

from datetime import datetime, date, timedelta
import json
from conexion import conectar
import spacy
from spacy.lang.es.stop_words import STOP_WORDS
from dateutil.relativedelta import relativedelta
from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for,current_app

from werkzeug.security import check_password_hash, generate_password_hash
import controllers.chatbot

import jinja2
from jinja2 import Environment



from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from fpdf import FPDF
import smtplib

#IMPORTS PARA GENERAR EL EXCEL
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,Font,Border,Side

from controllers.correo import enviar_usuario,enviar_correo_orden,enviar_correo_registro

bp = Blueprint('web', __name__)


nlp = spacy.load('es_core_news_sm')

# Palabras a excluir del conjunto stop_words
palabras_excluidas = ["tener","tienen"]

        # Remover palabras del conjunto stop_words
stop_words_personalizado = set(STOP_WORDS) - set(palabras_excluidas)

        # Asignar el nuevo conjunto stop_words personalizado al modelo de SpaCy
nlp.Defaults.stop_words = stop_words_personalizado

def generar_contrasena(longitud=8):
    if longitud < 4:
        raise ValueError("La longitud mínima recomendada es 4 caracteres.")

    letras = string.ascii_lowercase
    numeros = string.digits
    mayusculas = string.ascii_uppercase

    # Garantizamos al menos una letra minúscula, una mayúscula y un número
    contrasena = [
        random.choice(letras),
        random.choice(mayusculas),
        random.choice(numeros),
    ]

    # Rellenamos el resto de la contraseña con letras y números
    todos = letras + mayusculas + numeros
    contrasena += random.choices(todos, k=longitud - 3)

    # Mezclamos los caracteres
    random.shuffle(contrasena)

    return ''.join(contrasena)

def capturarHora():
    hi = datetime.now()
    return hi

@bp.route('/')
def home():

    # conn = conectar()
    # cursor = conn.cursor()
    # query = 'INSERT INTO credenciales (usuario,contrasena,Rol,cargo) VALUES (?,?,1,1)'
    # cursor.execute(query, ('JHOEL',generate_password_hash('123')))
    # conn.commit()
    # cursor.close()
    # conn.close()
    # conn = conectar()
    # cursor = conn.cursor()



    return render_template('web/index.html')

# INICIO DE LA TIENDA
@bp.route('/tienda')
def tienda():
    return render_template('web/tienda.html')


# INICIO DE LA CONSULTA CALENDARIO  
@bp.route('/consulta')
def consulta():
    
    return render_template('web/cliente.html',mascota = '',rol =session["rol"],nombre =session["nombre"], cons = '')
                            
# INICIO DE LA CONSULTA CALENDARIO  
@bp.route('/recuperar', methods=['POST'])
def recuperar():
    datos = request.form['datos']

    print(datos)
    conn = conectar()
            
    cursor = conn.cursor()
    query = 'select c.correo_cliente,cred.id_credencial,cred.usuario from credenciales as cred inner join cliente as c on cred.id_credencial = c.id_credencial inner join roles as r on cred.rol = r.cod_rol where (cred.usuario = ? or c.correo_cliente = ?) AND c.id_estado = 1'
    cursor.execute(query, (datos,datos))
    correo = cursor.fetchall()
    cursor.close()
    conn.close()
    if correo:

        contra = generar_contrasena(8)
        usuario = correo[0][2]

        conn = conectar()
        cursor = conn.cursor()
        query = 'UPDATE credenciales set contrasena = ? where id_credencial = ?'
        cursor.execute(query,(generate_password_hash(contra),correo[0][1]))
        conn.commit()
        cursor.close()
        conn.close()



        enviar_correo_registro(current_app,"Usuario Actualizado!!!",correo[0][0],'registro',usuario,contra)

        return 'Hecho'
    else:
        return 'No'
         
# LOGIN 

@bp.route('/login', methods=['POST'])
def login():
    if request.method == "POST":
        usuario = request.form['usuario']
        contraseña = request.form['pass']

        if usuario == "" or contraseña == "":
            return render_template('web/index.html')

        else:
            conn = conectar()
            cursor = conn.cursor()
            query = 'select c.num_cliente,c.nombres_cliente,c.apellidos_cliente,c.correo_cliente,c.telefono,cred.contrasena,r.nombre_rol from credenciales as cred inner join cliente as c on cred.id_credencial = c.id_credencial inner join roles as r on cred.rol = r.cod_rol where cred.usuario = ? AND c.id_estado = 1'
            cursor.execute(query, usuario)
            rows = cursor.fetchone()
            cursor.close()
            conn.close()
            print(rows)
            if rows:
                if len(rows) == 0 or not check_password_hash(rows[5], contraseña):
                    return render_template('web/index.html')
                else:
                    session['id'] = rows[0]
                    session['nombre'] = rows[1]
                    session["pass"] = contraseña
                    session['rol'] = rows[6]

                    if session['rol'] == 'ADMINISTRADOR' or session['rol'] == 'USUARIO' or session['rol'] == 'VETERINARIO':
                         return redirect('/sistema')  # Cambia esto por la ruta correcta de sistema.py
                    else:
                        #session['last_seen'] = datetime.now()
                         #aca traemos las consultas del cliente logeado
                    
                        
                        #receta = db1.execute('select * from DetalleReceta Where IdReceta = :rec',rec = historial[0]['Id_Receta'])
                        return render_template('web/cliente.html',mascota = '',rol =session["rol"],nombre =session["nombre"], cons = '')
                                

                    
            else:
               return render_template('web/index.html')
        print('fuera')                     
    return render_template('web/index.html')



# FIN LOGIN



# INICIO CONSULTAS
@bp.route('/traerCitas')
def traerCitas():
    print('web')
    conn = conectar()
    cursor = conn.cursor()
    query = "select a.cod_atencion,c.nombres_cliente,a.fecha_atencion,e.NombreEstado from atencion as a inner join cliente as c on a.num_cliente = c.num_cliente INNER JOIN estado as e on a.id_estado = e.id_estado where e.NombreEstado = 'AGENDADO' and c.num_cliente = ?"
    cursor.execute(query,(session['id']))
    agendas = cursor.fetchall()
    agenda= []
    for fila in agendas:
        fecha_inicio = fila[2]
        fecha_fin = fecha_inicio + timedelta(minutes=30)  # Suma 30 minutos a la hora de inicio

        agendas = {
            'numero': fila[0],
            'cliente': fila[1],
            'estado': fila[3],
            'fecha': fecha_inicio.strftime('%Y-%m-%d %H:%M:%S'),  
            'fechafin': fecha_fin.strftime('%Y-%m-%d %H:%M:%S') 
        }
        agenda.append(agendas)

    return ''+str(agenda)


#FIN CONSULTAS

# Buscador de la tienda
@bp.route('/buscarProducto', methods=['POST'])
def buscarProducto():

    if request.method == "POST":
        producto = request.form['producto']

        conn = conectar()
        cursor = conn.cursor()
        query = "select cod_producto,precio,Imagen,stock,nom_producto from producto where nom_producto like ? and Tienda = 'Si'"
        cursor.execute(query,(producto + '%'))
        productos = cursor.fetchall()
        print(productos)
        return render_template('web/otros/buscador_productos.html', productos=productos)
        
    else:
        return "No"
    
    

# Fin buscador de la tienda
# Llamado detalle del producto
@bp.route('/detalleProducto', methods=['POST'])
def detalleProducto():

    if request.method == "POST":
        num = request.form['num']

        conn = conectar()
        cursor = conn.cursor()
        query = 'select  cod_producto,precio,Imagen,stock,stock_critico,nom_producto from producto where cod_producto = ?'
        cursor.execute(query,(num))
        datos = cursor.fetchall()
        print(datos)
        if datos:

            return render_template('web/modales/modal_compra.html', producto=datos)
        else:
            return "Sin Datos"
    else:
        return "No"
    
    

# Fin buscador de la tienda


# RESTAR PRODUCTO DEL CARRITO
@bp.route('/restarProducto', methods=['POST'])
def restarProducto():

    if request.method == "POST":
        
        producto = request.form['producto']
        print(producto)

        conn = conectar()
        cursor = conn.cursor()
        query = 'select cod_producto,cantidad from detalle_carrito where cod_detalle = ?'
        cursor.execute(query,(producto))
        datos = cursor.fetchone()

        if datos[0] <= 1:
            return 'no'



        conn = conectar()
        cursor = conn.cursor()
        query = 'UPDATE detalle_carrito set cantidad -= 1 where cod_detalle = ?'
        cursor.execute(query,(producto))
        conn.commit()
        cursor.close()
        conn.close()

        conn = conectar()
        cursor = conn.cursor()
        query = 'select cantidad from detalle_carrito where cod_detalle = ?'
        cursor.execute(query,(producto))
        datos = cursor.fetchone()
        
        return str(datos[0])
    else:
        return "No"
    
    return render_template('web/otros/buscador_productos.html')

# Fin RESTA DE PRODUCTO

# RESTAR PRODUCTO DEL CARRITO
@bp.route('/sumarCantidad', methods=['POST'])
def sumarCantidad():

    if request.method == "POST":
        
        producto = request.form['producto']
        print(producto)

        conn = conectar()
        cursor = conn.cursor()
        query = 'select cod_producto,cantidad from detalle_carrito where cod_detalle = ?'
        cursor.execute(query,(producto))
        product = cursor.fetchone()

        cantidad = product[1]

        #AQUI REVISAMOS EL STOCK DEL MATERIAL QUE ESTA EN EL CARRITO
        conn = conectar()
        cursor = conn.cursor()
        query = 'select stock from producto where cod_producto = ?'
        cursor.execute(query,(product[0]))
        stock = cursor.fetchone()

        if cantidad + 1 <= stock[0]:


            conn = conectar()
            cursor = conn.cursor()
            query = 'UPDATE detalle_carrito set cantidad += 1 where cod_detalle = ?'
            cursor.execute(query,producto)
            conn.commit()
            cursor.close()
            conn.close()

            conn = conectar()
            cursor = conn.cursor()
            query = 'select cantidad from detalle_carrito where cod_detalle = ?'
            cursor.execute(query,(producto))
            datos = cursor.fetchone()
            
            return str(datos[0])
        else:
            return 'Sin Stock'
    else:
        return "No"
    
    return render_template('web/otros/buscador_productos.html')
# Fin RESTA DE PRODUCTO
# MANDAR ORDEN DE COMPRA
@bp.route('/mandarSalida', methods=["POST"])
def mandarSalida():
    num = request.form['id']
    tempPdfFilePath = r'C:\Users\JHOEL SEQUEIRA\Documents\GitHub\monografia\ordenCompra\Orden.pdf'

    conn = conectar()
    cursor = conn.cursor()

    # Consulta SQL para obtener los datos del ticket
    query = """
    SELECT sm.NumSalida, t.Tramportista, c.Placa, r.Rastra, con.NombreConductor, p.NombreProveedor, 
    m.NombreMaterial, sm.PesoBruto, sm.PesoTara, sm.PesoTara2, sm.PesoNeto, sm.Observacion, sm.FechaIngreso, sm.FechaSalida 
    FROM SalidaMaterial as sm 
    INNER JOIN Camiones as c ON sm.NumCamion = c.NumCamion 
    INNER JOIN Tramportistas as t ON sm.NumTramportista = t.NumTramportista 
    INNER JOIN Rastras as r ON r.NumRastra = sm.NumRastra
    INNER JOIN Materiales as m ON sm.NumMaterial = m.NumMaterial 
    INNER JOIN Proveedores as p ON sm.NumProveedor = p.NumProveedor 
    INNER JOIN Conductores as con ON sm.NumConductor = con.NumConductor 
    WHERE sm.NumSalida = ?
    """

    cursor.execute(query, (num,))
    ticket = cursor.fetchone()

    fecha_ingreso = ticket[12].strftime("%Y-%m-%d | %I:%M:%S %p") if isinstance(ticket[12], datetime) else str(ticket[12])
    fecha_salida = ticket[13].strftime("%Y-%m-%d | %I:%M:%S %p") if isinstance(ticket[13], datetime) else str(ticket[13])
    fecha = ticket[13].strftime("%Y-%m-%d") if isinstance(ticket[13], datetime) else str(ticket[13])

    pdf = FPDF('P', 'mm', (101.6, 175))
    pdf.set_margins(5.5, 5.5, 5.5)
    pdf.set_display_mode(zoom=100, layout='continuous')
    pdf.add_page()

    pdf.set_font('Arial', 'B', 30)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Arial', 'B', 14)
    pdf.multi_cell(0, 5, 'Compañía Recicladora de Nicaragua', 0, "C")
    pdf.ln(1)
    
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(25, 10, 'No de Boleta: ')
    pdf.set_font('Arial', '', 7.5)
    pdf.cell(30, 10, str(ticket[0]))
    pdf.ln(6)
    
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(25, 10, 'Fecha: ')
    pdf.set_font('Arial', 'B', 7.5)
    pdf.multi_cell(23, 10, fecha)
    pdf.ln(4)
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(20, 10, 'Salida de Material ')
    pdf.ln(8)  

    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(25, 10, 'Placa: ')
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(20, 10, ticket[3])
    pdf.ln(6)
    
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(25, 10, 'Conductor: ')
    pdf.set_font('Arial', '', 7.5)
    pdf.cell(25, 10, ticket[4].upper())
    pdf.ln(9)
    
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(25, 10, 'Cliente: ')
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(30, 10, ticket[5].upper())
    pdf.ln(9)

    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(25, 10, 'Material: ')
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(20, 10, ticket[6])
    pdf.ln(9)
    
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(25, 10, 'Peso Total: ')
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(25, 10, "{:,.2f}".format(float(ticket[7])) + ' lb')
    pdf.ln(6)

    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(25, 10, 'Peso Tara: ')
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(25, 10, "{:,.2f}".format(float(ticket[8])) + ' lb')
    pdf.ln(6)

    # pdf.set_font('Arial', 'B', 7.5)
    # pdf.cell(30, 10, 'Peso Tara Adicional: ')
    # pdf.set_font('Arial', '', 7.5)
    # pdf.cell(20, 10,"{:,.2f}".format(float(ticket[9]) - float(ticket[7])) + 'lb')
    pdf.ln(8)

    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(30, 10, 'Peso Bruto: ')
    pdf.set_font('Arial', '', 7.5)
    pdf.cell(20, 10, "{:,.2f}".format(float(ticket[10])) + ' lb')
    kg = float(ticket[10]) / 2.2046
    pdf.ln(6)
    

    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(30, 10, '')
    pdf.cell(20, 10, "{:,.2f}".format(kg) + ' kg')
    pdf.ln(6)



    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(30, 10, 'Fecha y Hora Ingreso: ')
    pdf.set_font('Arial', '', 7.5)
    pdf.cell(30, 10, fecha_ingreso)
    pdf.ln(4)



    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(30, 10, 'Fecha y Hora Salida: ')
    pdf.set_font('Arial', '', 7.5)
    pdf.cell(30, 10, fecha_salida)
    pdf.ln(4)

    

   

    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(30, 10, 'Observacion: ')
    pdf.ln(6)
    pdf.set_font('Arial', 'B', 7.5)
    pdf.cell(180, 10, ticket[11].upper())
    pdf.ln(6)

    # pdf.set_font('Arial', 'B', 7.5)
    # pdf.cell(30, 10, ticket[12],0, 0, 'C')
    # pdf.ln(3)


    x = 35
    y = 110
    width = 30
    height = 0 
    imagePath = 'static/img/validado.png'

    pdf.image(imagePath, x, y, width, height)

   
    pdf.cell(30, 10, '___________________',0, 0, 'C')
    pdf.cell(30, 10, '___________________',0, 0, 'C')
    pdf.cell(30, 10, '___________________',0, 0, 'C')
    pdf.ln(6)
    pdf.cell(30, 10, 'Digitador',0, 0, 'C')
    pdf.cell(30, 10, 'Verificador',0, 0, 'C')
    pdf.cell(30, 10, 'Proveedor',0, 0, 'C')

    # Guardar el PDF como un archivo temporal
    pdf.output(tempPdfFilePath, 'F')

    # Configurar los detalles del correo
    from_email = 'bascula@crn.com.ni'
    to_email = ['maycol.gonzalez@crn.com.ni', 'david.garache@crn.com.ni', 'aurelio.larios@crn.com.ni', 'exportaciones@crn.com.ni']
    # to_email = ['it@crn.com.ni']
    cc_emails = ['bryan.oviedo@crn.com.ni']
    # cc_emails = ['it@crn.com.ni']
    subject = ticket[3]
    body = f"Se adjunta el checklist validado de la unidad: <b>{ticket[3]}</b><br> Material: <b>{ticket[6]}</b>"

    # Configurar el servidor SMTP
    smtp_server = 'smtp.office365.com'
    smtp_port = 587
    smtp_user = 'bascula@crn.com.ni'
    smtp_password = 'Cbasscula23'

    try:
         # Crear el mensaje de correo
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = ', '.join(to_email)  # Corregido: Convertir la lista de destinatarios a cadena
        msg['Cc'] = ', '.join(cc_emails)
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))

        # Adjuntar el archivo PDF
        with open(tempPdfFilePath, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(tempPdfFilePath)}')
            msg.attach(part)

        # Enviar el correo
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            text = msg.as_string()
            server.sendmail(from_email, to_email, text)

        # Eliminar el archivo temporal después de enviar el correo
        os.remove(tempPdfFilePath)
        print('correo enviado')
        return 'Correo enviado con éxito'

    except Exception as e:
        print( {str(e)})
        return f'Error al enviar el correo: {str(e)}'


# FIN ENVIO ORDEN DE COMPRA

# RESTAR PRODUCTO DEL CARRITO
@bp.route('/borrarProducto', methods=['POST'])
def borrarProducto():

    if request.method == "POST":
        
        producto = request.form['producto']
        print(producto)

        conn = conectar()
        cursor = conn.cursor()
        query = 'delete from detalle_carrito where cod_detalle = ?'
        cursor.execute(query,(producto))
        conn.commit()
        cursor.close()
        conn.close()

        return 'hecho'
    else:
        return "No"
    
    return render_template('web/otros/buscador_productos.html')

# Fin RESTA DE PRODUCTO



# Llamado del Carrito total por Usuario
@bp.route('/carritoTotal', methods=['POST'])
def carritoTotal():

    if request.method == "POST":

        conn = conectar()
        cursor = conn.cursor()
        query = 'SELECT cc.id_carrito,p.Imagen, p.nom_producto, cc.cantidad, p.precio, e.NombreEstado FROM carrito_compra AS cc INNER JOIN producto AS p ON cc.cod_producto = p.cod_producto INNER JOIN cliente AS c ON cc.num_cliente = c.num_cliente INNER JOIN estado AS e ON cc.id_estado = e.id_estado where cc.num_cliente = ? GROUP BY p.nom_producto, cc.id_carrito, cc.cantidad, p.precio, e.NombreEstado, p.Imagen;'
        cursor.execute(query,(session['id']))
        datos = cursor.fetchall()
        print(datos)
        if datos:

            return render_template('web/modales/modal_carrito.html', producto=datos)
        else:
            return "Sin Datos"
    else:
        return "No"
    
    return render_template('web/buscador_productos.html')


@bp.route('/carrito')
def carrito():


        return render_template('web/carrito.html')

@bp.route('/cargarProductosCarrito', methods=['POST'])
def cargarProductosCarrito():

        conn = conectar()
        cursor = conn.cursor()
        query = 'SELECT dc.cod_detalle,p.Imagen, p.nom_producto, dc.cantidad, p.precio, e.NombreEstado FROM carrito_compra AS cc inner join detalle_carrito as dc on cc.id_carrito = dc.cod_carrito INNER JOIN producto AS p ON dc.cod_producto = p.cod_producto INNER JOIN cliente AS c ON cc.num_cliente = c.num_cliente INNER JOIN estado AS e ON cc.id_estado = e.id_estado where cc.num_cliente = ? and cc.id_estado = 2 GROUP BY p.nom_producto, dc.cod_detalle, dc.cantidad, p.precio, e.NombreEstado, p.Imagen;'
        cursor.execute(query,(session['id']))
        datos = cursor.fetchall()
        print(datos)
        if not datos:
            datos = ''

        return render_template('web/otros/traer_productos.html', producto=datos)


@bp.route('/generarOrdenCompra', methods=['POST'])
def generarOrdenCompra():
        FechaActual = capturarHora()  
        conn = conectar()
        cursor = conn.cursor()
        query = 'SELECT cc.id_carrito from carrito_compra as cc inner join estado as e on cc.id_estado = e.id_estado where cc.num_cliente = ? and cc.id_estado = 2 '
        cursor.execute(query,(session['id']))
        carrito = cursor.fetchone()


        conn = conectar()
        cursor = conn.cursor()
        query = 'INSERT INTO venta (fecha_venta,Total,num_cliente,vendedor,cod_estado,id_carrito) VALUES (?,?,?,?,15,?)'
        cursor.execute(query, (FechaActual,0,session['id'],1,carrito[0]))
        conn.commit()
        cursor.close()
        conn.close()

        conn = conectar()
        cursor = conn.cursor()
        query = 'SELECT top 1 v.cod_venta,c.correo_cliente from venta as v inner join cliente as c on v.num_cliente = c.num_cliente where v.num_cliente = ? order by v.cod_venta desc '
        cursor.execute(query,(session['id']))
        id = cursor.fetchone()

        enviar_correo_orden(current_app,"Bienvenido a Nuestra Familia",id[1],'orden',id[0])
        
        return str(id[0])







@bp.route('/seguimiento', methods=['POST'])
def seguimiento():

        conn = conectar()
        cursor = conn.cursor()
        query = 'SELECT cc.id_carrito,e.NombreEstado as estado from carrito_compra as cc inner join estado as e on cc.id_estado = e.id_estado where cc.num_cliente = ? and cc.id_estado != 2 '
        cursor.execute(query,(session['id']))
        carrito = cursor.fetchall()

        return render_template('web/otros/seguimiento.html', producto=carrito)

@bp.route('/cantidadArticulos', methods=['POST'])
def cantidadArticulos():

        conn = conectar()
        cursor = conn.cursor()
        query = 'SELECT COUNT(DISTINCT p.cod_producto) AS total_productos FROM carrito_compra AS cc INNER JOIN detalle_carrito as dc on cc.id_carrito = dc.cod_carrito inner join producto AS p ON dc.cod_producto = p.cod_producto INNER JOIN cliente AS c ON cc.num_cliente = c.num_cliente INNER JOIN estado AS e ON cc.id_estado = e.id_estado WHERE cc.num_cliente = ? and cc.id_estado = 2;'
        cursor.execute(query,(session['id']))
        datos = cursor.fetchone()
        print(datos)
        if not datos[0]:
            datos = 0
        else:
            datos = datos[0]
        return render_template('web/otros/cantidad_articulos.html', producto=datos)
    
    

# Fin buscador de la tienda

# Guardar carrito del cliente
@bp.route('/guardarCarrito', methods=['POST'])
def guardarCarrito():

    if request.method == "POST":
        if session:
            producto = request.form['producto']
            cantidad = request.form['cantidad']


            conn = conectar()
            cursor = conn.cursor()
            query = 'select * from carrito_compra where num_cliente = ? and id_estado = 2'
            cursor.execute(query, (session['id']))
            carrito = cursor.fetchone()
            cursor.close()
            conn.close()

            print(carrito)

            if carrito:


                conn = conectar()
                cursor = conn.cursor()
                query = 'select * from detalle_carrito where cod_producto = ? and cod_carrito = ?'
                cursor.execute(query, (producto,carrito[0]))
                detallecarrito = cursor.fetchone()
                cursor.close()
                conn.close()

                print(detallecarrito)
                if detallecarrito:

                    conn = conectar()
                    cursor = conn.cursor()
                    query = 'UPDATE detalle_carrito set cantidad += ? where cod_detalle = ? and cod_producto = ?'
                    cursor.execute(query,(cantidad,detallecarrito[0],producto))
                    conn.commit()
                    cursor.close()
                    conn.close()
                else:
                    print('else')
                    print()
                    conn = conectar()
                    cursor = conn.cursor()
                    query = 'INSERT INTO detalle_carrito (cod_producto,cantidad,cod_carrito,precio_venta) VALUES (?,?,?,0)'
                    cursor.execute(query,(producto,cantidad,carrito[0]))
                    conn.commit()
                    cursor.close()
                    conn.close()

            else:

                #INSERTAMOS EL CARRITO
                conn = conectar()
                cursor = conn.cursor()
                query = 'INSERT INTO carrito_compra (num_cliente,id_estado) VALUES (?,2)'
                cursor.execute(query,(session['id']))
                conn.commit()
                cursor.close()
                conn.close()

                #SELECCIONAMOS EL ULTIMO CARRITO AGREGADO
                conn = conectar()
                cursor = conn.cursor()
                query = 'select TOP 1 * from carrito_compra where num_cliente = ? and id_estado = 2 order by id_carrito desc'
                cursor.execute(query, (session['id']))
                car = cursor.fetchone()
                cursor.close()
                conn.close()

                print(car[0])
                #INSERTAMOS EL DETALLE DEL CARRITO
                
                conn = conectar()
                cursor = conn.cursor()
                query = 'INSERT INTO detalle_carrito (cod_producto,cantidad,cod_carrito) VALUES (?,?,?)'
                cursor.execute(query,(producto,cantidad,car[0]))
                conn.commit()
                cursor.close()
                conn.close()
            
            return 'HECHO'
        else:
            return 'Sin Sesion'
    else:
        return "No"
    
    

# Fin guardar carrito de la tienda

@bp.route('/comprar', methods=['POST'])
def comprar():
    if request.method == "POST":
        usuario = request.form['usuario']
        contraseña = request.form['pass']

        if usuario == "" or contraseña == "":
            return render_template('login.html', errorlogin=1)
        else:
            conn = conectar()
            cursor = conn.cursor()
            query = 'select c.num_cliente,c.nombres_cliente,c.apellidos_cliente,c.correo_cliente,c.telefono,cred.contrasena,r.nombre_rol from credenciales as cred inner join cliente as c on cred.id_credencial = c.id_credencial inner join roles as r on cred.rol = r.cod_rol where cred.usuario = ? AND c.id_estado = 1'
            cursor.execute(query, usuario)
            rows = cursor.fetchone()
            cursor.close()
            conn.close()
            print(rows)
            if rows:
                if len(rows) == 0 or not check_password_hash(rows[5], contraseña):
                    return 'error'
                else:
                    #session['last_seen'] = datetime.now()
                    session['id'] = rows[0]
                    session['nombre'] = rows[1]
                    session["pass"] = contraseña
                    session['rol'] = rows[6]
                return 'exito'
            else:
               return 'error'
    return 'error'

# CargarCarrito Numero

@bp.route('/cargarCarrito', methods=['POST'])
def cargarCarrito():

    if request.method == "POST":
        if session:

            conn = conectar()
            cursor = conn.cursor()
            query = 'select count(*) from carrito_compra  as cc inner join detalle_carrito as dc on cc. id_carrito = dc.cod_carrito where cc.num_cliente = ? and cc.id_estado = 2'
            cursor.execute(query, session['id'])
            rows = cursor.fetchone()
            cursor.close()
            conn.close()
            
            return str(rows[0])
        else:
            return 'Sin Sesion'
    else:
        return "No"

# TAbla de precios en el carrito de la persona
@bp.route('/cargarTabla', methods=["GET", "POST"])
def cargarTabla():

    conn = conectar()
    cursor = conn.cursor()
    query = 'SELECT dc.cod_detalle,p.Imagen, p.nom_producto, dc.cantidad, p.precio, e.NombreEstado FROM carrito_compra AS cc inner join detalle_carrito as dc on cc.id_carrito = dc.cod_carrito INNER JOIN producto AS p ON dc.cod_producto = p.cod_producto INNER JOIN cliente AS c ON cc.num_cliente = c.num_cliente INNER JOIN estado AS e ON cc.id_estado = e.id_estado where cc.num_cliente = ? and cc.id_estado = 2 GROUP BY p.nom_producto, dc.cod_detalle, dc.cantidad, p.precio, e.NombreEstado, p.Imagen;'
    cursor.execute(query,(session['id']))
    datos = cursor.fetchall()


    conn = conectar()
    cursor = conn.cursor()
    query = 'SELECT SUM(dc.cantidad * p.precio) AS total_factura FROM carrito_compra AS cc inner join detalle_carrito as dc on cc.id_carrito = dc.cod_carrito INNER JOIN producto AS p ON dc.cod_producto = p.cod_producto INNER JOIN cliente AS c ON cc.num_cliente = c.num_cliente WHERE cc.num_cliente = ?'
    cursor.execute(query,(session['id']))
    total = cursor.fetchone()
    
    return render_template('web/tablas/tabla_precios.html', productos = datos,total = total[0])


#FIN TABLA






# FIN DE LA TIENDA

# MODULO DE CHATBOT
@bp.route('/recibirMensaje',methods=["GET", "POST"])
def recibirMensaje():
    if request.method == 'POST':

        solicitud = request.form['mensaje']
        doc = nlp(solicitud.lower())
        # LINEA CON LAS STOP_WORDS PRESENTA PROBLEMAS CON ALGUNOS VERBOS
        #lemmatized_tokens = [token.lemma_ for token in doc if token.is_alpha and not token.is_stop]
        #LINEA SIN LOS STOP WORDS
        lemmatized_tokens = [token.lemma_ for token in doc if token.is_alpha ]

        filtered_tokens = [token for token in lemmatized_tokens if token.isalnum()]
        print("PRIMER FILTRO: ", filtered_tokens)
        if session.get("id"):
            resultado = controllers.chatbot.procesar_entrada(filtered_tokens, session["id"], current_app)
            return resultado
        else:
            return "Su usuario no fue encontrado, usted podrá hacer consultas sencillas como horarios, precios, servicios, etc."




# FIN CHATBOT
# BUSCAR USUARIOS EXISTENTES
@bp.route('/buscarusu', methods =["POST","GET"])
def buscarusu():
    if request.method == "POST":
        usuario = request.form['usuario']
        
        conn = conectar()
        cursor = conn.cursor()
        query = 'SELECT * from Credenciales WHERE Usuario = ?'
        cursor.execute(query,(usuario))
        existente = cursor.fetchone()
        if existente:
            return "ya hay"
        else:
            return "no existe"

@bp.route('/buscarCorreo', methods =["POST","GET"])
def buscarCorreo():
    if request.method == "POST":
        correo = request.form['correo']
        
        conn = conectar()
        cursor = conn.cursor()
        query = 'SELECT * from cliente WHERE correo_cliente like ?'
        cursor.execute(query,(correo+'%'))
        existente = cursor.fetchone()
        if existente:
            return "ya hay"
        else:
            return "no existe"

# FIN BUSCAR USUARIOS
#  AÑADIR USUARIO
@bp.route('/nuevou',methods=["GET", "POST"])
def nuevou():
    if request.method == "POST":
        nombres = request.form['nombres']
        apellidos = request.form['apellidos']
        direccion = request.form['direccion']
        telefono = request.form['telefono']
        celular = request.form['celular']
        correo = request.form['correo']
        usuario = request.form['loginUser']
        contraseña = request.form['loginPassword']
        flag = request.form['flag']
        print(flag)
        print("ENTROOOO")
        if flag:
            rol = request.form['rol']
            # verificacion = db1.execute('SELECT * from credenciales Where Usuario = :u',u = usuario)
            # db1.execute("INSERT INTO Credenciales VALUES(NULL,:usu,:passw,:rol)",usu = usuario,passw = generate_password_hash(contraseña),rol = rol)
            
            conn = conectar()
            cursor = conn.cursor()
            query = 'INSERT INTO credenciales (usuario,contrasena,rol,cargo) VALUES (?,?,?,?)'
            cursor.execute(query, (usuario,generate_password_hash(contraseña),rol,4))
            conn.commit()
            cursor.close()
            
            # credenciales = db1.execute('select Id_Credenciales from Credenciales  order by Id_Credenciales desc limit 1')


            conn = conectar()
            cursor = conn.cursor()
            query = 'SELECT id_credencial from credenciales order by id_credencial desc'
            cursor.execute(query)
            existente = cursor.fetchone()



            # db1.execute("INSERT INTO cliente VALUES(NULL,:name,:lastna,:tel,:cel,:corr,:dir,1,:cred,null)",
            #             name=nombres , lastna=apellidos , tel = telefono,cel = celular,corr = correo,dir=direccion,cred = credenciales[0]['Id_Credenciales'])
            
            conn = conectar()
            cursor = conn.cursor()
            query = 'INSERT INTO cliente (nombres_cliente,id_estado,apellidos_cliente,direccion_cliente,correo_cliente,id_credencial,telefono) VALUES (?,1,?,?,?,?,?)'
            cursor.execute(query, (nombres,apellidos,direccion,correo,existente[0],telefono))
            conn.commit()
            cursor.close()


            return "yes"
        else:
            conn = conectar()
            cursor = conn.cursor()
            query = 'INSERT INTO credenciales (usuario,contrasena,rol,cargo) VALUES (?,?,?,?)'
            cursor.execute(query, (usuario,generate_password_hash(contraseña),4,4))
            conn.commit()
            cursor.close()
            
            # credenciales = db1.execute('select Id_Credenciales from Credenciales  order by Id_Credenciales desc limit 1')


            conn = conectar()
            cursor = conn.cursor()
            query = 'SELECT id_credencial from credenciales order by id_credencial desc'
            cursor.execute(query)
            existente = cursor.fetchone()



            # db1.execute("INSERT INTO cliente VALUES(NULL,:name,:lastna,:tel,:cel,:corr,:dir,1,:cred,null)",
            #             name=nombres , lastna=apellidos , tel = telefono,cel = celular,corr = correo,dir=direccion,cred = credenciales[0]['Id_Credenciales'])
            
            conn = conectar()
            cursor = conn.cursor()
            query = 'INSERT INTO cliente (nombres_cliente,id_estado,apellidos_cliente,direccion_cliente,correo_cliente,id_credencial,telefono) VALUES (?,1,?,?,?,?,?)'
            cursor.execute(query, (nombres,apellidos,direccion,correo,existente[0],telefono))
            conn.commit()
            cursor.close()

            return 'yes'
    else:
        return render_template('nuevous.html')
    

@bp.route('/correo',methods=["GET", "POST"])
def correo():
    if request.method == "POST":
        flag = request.form['flag']
        print(flag)
        if flag == "usuario":
            nombres = request.form['nombres']
            apellidos = request.form['apellidos']
            correoc = request.form['correo']
            usuario = request.form['loginUser']
            contraseña = request.form['loginPassword']
            
            enviar_usuario(current_app,"Bienvenido a Nuestra Familia",correoc,'usuario',usuario,contraseña)
            return "done"
        
       
    else:
        return render_template('nuevous.html')

# FIN AÑADIR USUARIO
@bp.route('/servicios')
def servicios():
    return render_template('web/servicios.html')

@bp.route('/clientes')
def clientes():
    return render_template('web/clientes.html')

@bp.route('/cerrarSesion')
def cerrarSesion():
    session.clear()
    return render_template('web/index.html')
